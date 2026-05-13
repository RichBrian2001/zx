from __future__ import annotations

import json
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFilter, ImageFont


ROOT = Path(__file__).resolve().parents[1]
UPLOAD_ROOT = ROOT / "demo-assets" / "uploadPath"
RUNTIME_UPLOAD_ROOT = UPLOAD_ROOT / "upload"
PLANT_DIR = RUNTIME_UPLOAD_ROOT / "plant"
ARTICLE_DIR = RUNTIME_UPLOAD_ROOT / "article"
IMPORT_DIR = RUNTIME_UPLOAD_ROOT / "import"
MANIFEST_PATH = ROOT / "demo-assets" / "asset-manifest.json"


PLANT_ASSETS = [
    {"plant_id": 1001, "name": "绿萝", "latin": "Epipremnum", "palette": ("#e6f4db", "#6ca968", "#2f6b35"), "variant": "vine"},
    {"plant_id": 1002, "name": "吊兰", "latin": "Spider Plant", "palette": ("#eef5de", "#89b56a", "#456b32"), "variant": "spider"},
    {"plant_id": 1003, "name": "虎皮兰", "latin": "Sansevieria", "palette": ("#ebe9d6", "#7a9960", "#3f5a2d"), "variant": "snake"},
    {"plant_id": 1004, "name": "龟背竹", "latin": "Monstera", "palette": ("#e1f0de", "#5a9b6a", "#245d45"), "variant": "monstera"},
    {"plant_id": 1005, "name": "发财树", "latin": "Pachira", "palette": ("#efe8d9", "#7ba15c", "#4d6930"), "variant": "money_tree"},
    {"plant_id": 1006, "name": "白掌", "latin": "Peace Lily", "palette": ("#f4efe4", "#77ab72", "#39693c"), "variant": "peace_lily"},
    {"plant_id": 1007, "name": "常春藤", "latin": "English Ivy", "palette": ("#dfeee2", "#729f63", "#2d5b35"), "variant": "ivy"},
    {"plant_id": 1008, "name": "橡皮树", "latin": "Rubber Tree", "palette": ("#f2eadf", "#5f8f5e", "#274c33"), "variant": "rubber"},
    {"plant_id": 1009, "name": "文竹", "latin": "Asparagus Fern", "palette": ("#edf3e3", "#8aa86d", "#466337"), "variant": "fern"},
    {"plant_id": 1010, "name": "茉莉花", "latin": "Jasmine", "palette": ("#f7f2e7", "#6eaa65", "#2d6d40"), "variant": "jasmine"},
    {"plant_id": 1011, "name": "薄荷", "latin": "Mint", "palette": ("#e2f3e8", "#7db77a", "#2f7255"), "variant": "mint"},
    {"plant_id": 1012, "name": "玉露", "latin": "Haworthia", "palette": ("#e8f3f4", "#7eb4a2", "#2f6b63"), "variant": "succulent"},
    {"plant_id": 1013, "name": "金钱树", "latin": "Zamioculcas", "palette": ("#eef1df", "#88a268", "#3f5e32"), "variant": "zz"},
    {"plant_id": 1014, "name": "袖珍椰子", "latin": "Parlor Palm", "palette": ("#e7f0e4", "#7aaa72", "#325c3d"), "variant": "palm"},
    {"plant_id": 1015, "name": "豆瓣绿", "latin": "Peperomia", "palette": ("#edf1e6", "#86ad73", "#45683d"), "variant": "peperomia"},
    {"plant_id": 1016, "name": "鸟巢蕨", "latin": "Bird's Nest Fern", "palette": ("#edf5e4", "#78aa70", "#3e6a45"), "variant": "fern"},
    {"plant_id": 1017, "name": "迷迭香", "latin": "Rosemary", "palette": ("#eef3ea", "#7fa98d", "#406256"), "variant": "rosemary"},
    {"plant_id": 1018, "name": "琴叶榕", "latin": "Fiddle Leaf Fig", "palette": ("#efe9df", "#6f9b64", "#2e5735"), "variant": "fiddle"},
    {"plant_id": 1019, "name": "孔雀竹芋", "latin": "Calathea", "palette": ("#edf2e8", "#8cab76", "#4d6846"), "variant": "fern"},
    {"plant_id": 1020, "name": "散尾葵", "latin": "Areca Palm", "palette": ("#edf3df", "#83ad6b", "#3d6541"), "variant": "palm"},
    {"plant_id": 1021, "name": "富贵竹", "latin": "Lucky Bamboo", "palette": ("#eef4e6", "#8cb06d", "#456b3a"), "variant": "zz"},
    {"plant_id": 1022, "name": "鹿角蕨", "latin": "Staghorn Fern", "palette": ("#e7f0e4", "#7aa472", "#3d6748"), "variant": "fern"},
    {"plant_id": 1023, "name": "合果芋", "latin": "Syngonium", "palette": ("#edf1e4", "#87ad7a", "#476644"), "variant": "vine"},
    {"plant_id": 1024, "name": "冷水花", "latin": "Aluminum Plant", "palette": ("#edf1ee", "#8aa79a", "#4d6762"), "variant": "peperomia"},
    {"plant_id": 1025, "name": "姬龟背竹", "latin": "Mini Monstera", "palette": ("#e4efe3", "#6aa36f", "#2f6347"), "variant": "monstera"},
    {"plant_id": 1026, "name": "芦荟", "latin": "Aloe Vera", "palette": ("#eaf3e8", "#8bb17b", "#466a3f"), "variant": "succulent"},
    {"plant_id": 1027, "name": "铜钱草", "latin": "Pennywort", "palette": ("#eff4e6", "#86b379", "#447049"), "variant": "peperomia"},
    {"plant_id": 1028, "name": "长寿花", "latin": "Kalanchoe", "palette": ("#f5eee2", "#91b06f", "#5f7040"), "variant": "jasmine"},
    {"plant_id": 1029, "name": "四季海棠", "latin": "Begonia", "palette": ("#f5ece3", "#98af79", "#5d6d45"), "variant": "jasmine"},
    {"plant_id": 1030, "name": "天竺葵", "latin": "Pelargonium", "palette": ("#f4ece1", "#8eae78", "#566c43"), "variant": "jasmine"},
    {"plant_id": 1031, "name": "口红吊兰", "latin": "Lipstick Plant", "palette": ("#e5efe5", "#7ea56e", "#3f6944"), "variant": "ivy"},
    {"plant_id": 1032, "name": "球兰", "latin": "Hoya", "palette": ("#eff2e8", "#88aa78", "#476447"), "variant": "vine"},
    {"plant_id": 1033, "name": "空气凤梨", "latin": "Tillandsia", "palette": ("#edf2ef", "#8cae9c", "#496867"), "variant": "succulent"},
    {"plant_id": 1034, "name": "网纹草", "latin": "Fittonia", "palette": ("#f0f2ea", "#93ab7d", "#586a4e"), "variant": "peperomia"},
    {"plant_id": 1035, "name": "铁线蕨", "latin": "Maidenhair Fern", "palette": ("#edf1ea", "#86a27b", "#476348"), "variant": "fern"},
    {"plant_id": 1036, "name": "龙血树", "latin": "Dracaena Marginata", "palette": ("#eee7df", "#7d9d6e", "#41593b"), "variant": "rubber"},
    {"plant_id": 1037, "name": "巴西木", "latin": "Corn Plant", "palette": ("#efe8de", "#83a06e", "#44603d"), "variant": "money_tree"},
    {"plant_id": 1038, "name": "春羽", "latin": "Philodendron Selloum", "palette": ("#e5f0df", "#6fa46d", "#346546"), "variant": "monstera"},
    {"plant_id": 1039, "name": "绿宝石喜林芋", "latin": "Green Philodendron", "palette": ("#e7efe0", "#79a06d", "#345d3d"), "variant": "rubber"},
    {"plant_id": 1040, "name": "酒瓶兰", "latin": "Ponytail Palm", "palette": ("#f1e8df", "#8aa06f", "#55623d"), "variant": "palm"},
    {"plant_id": 1041, "name": "栀子花", "latin": "Gardenia", "palette": ("#f7f1e7", "#8bad71", "#476744"), "variant": "jasmine"},
    {"plant_id": 1042, "name": "薰衣草", "latin": "Lavender", "palette": ("#efeaf1", "#9aa6b8", "#5f6e83"), "variant": "rosemary"},
    {"plant_id": 1043, "name": "百里香", "latin": "Thyme", "palette": ("#f0f2e8", "#93a180", "#59664b"), "variant": "rosemary"},
    {"plant_id": 1044, "name": "罗勒", "latin": "Basil", "palette": ("#e9f2e5", "#7fac6f", "#3f6b42"), "variant": "mint"},
    {"plant_id": 1045, "name": "柠檬香蜂草", "latin": "Lemon Balm", "palette": ("#eef4e8", "#8bb37c", "#4d724c"), "variant": "mint"},
    {"plant_id": 1046, "name": "非洲堇", "latin": "African Violet", "palette": ("#f1ecea", "#9d8fa7", "#665873"), "variant": "jasmine"},
    {"plant_id": 1047, "name": "银皇后", "latin": "Aglaonema", "palette": ("#eff2ea", "#9caf90", "#5d7059"), "variant": "rubber"},
]

ARTICLE_ASSETS = [
    {"slug": "bedroom", "title": "北向卧室如何选第一盆绿植", "subtitle": "弱光 · 小空间 · 低养护", "palette": ("#f2eee6", "#c6d7c3", "#8ca48c"), "scene": "bedroom"},
    {"slug": "office", "title": "办公室绿植 5 日养护法", "subtitle": "工位环境 · 低频浇水", "palette": ("#edf1f5", "#d7e2d4", "#8aa18b"), "scene": "office"},
    {"slug": "balcony", "title": "阳台花香植物入门清单", "subtitle": "日照充足 · 花香体验", "palette": ("#f8efe1", "#d9e5cb", "#99a771"), "scene": "balcony"},
    {"slug": "pet", "title": "猫咪家庭绿植避坑清单", "subtitle": "宠物友好 · 摆放安全", "palette": ("#f2ede9", "#d8e4d4", "#7e9e83"), "scene": "pet"},
    {"slug": "office-lowlight", "title": "低光办公室别只盯绿萝", "subtitle": "弱光工位 · 多样候选", "palette": ("#ecefe7", "#c8d2c3", "#7e9480"), "scene": "office_lowlight"},
    {"slug": "living", "title": "客厅大叶植物陈设的 3 个层次位", "subtitle": "主景 · 过渡位 · 焦点位", "palette": ("#efe8e0", "#d8d9cd", "#8f9f7d"), "scene": "living"},
    {"slug": "desktop", "title": "桌面小体量植物组合指南", "subtitle": "书桌 · 床头柜 · 工位", "palette": ("#f2efe7", "#dbe3d0", "#8fa481"), "scene": "desktop"},
    {"slug": "aroma", "title": "迷迭香、茉莉和薄荷怎么选", "subtitle": "花香 · 草本香 · 取用感", "palette": ("#f3eee4", "#d8e1cf", "#91a175"), "scene": "aroma"},
]

PLANT_TEMPLATE_HEADERS = [
    "分类ID", "植物名称", "植物编码", "别名", "封面图路径", "摘要", "难度等级", "价格等级", "适用场景", "备注"
]

PLANT_TEMPLATE_ROWS = [
    [102, "龟背竹", "monstera_deliciosa", "蓬莱蕉", "/profile/upload/plant/1004.jpg", "客厅大叶主景植物", 2, 2, "客厅落地、会客区、休闲角", "客厅陈设样例"],
    [105, "金钱树", "zamioculcas_zamiifolia", "雪铁芋", "/profile/upload/plant/1013.jpg", "低光办公样例植物", 1, 2, "办公室、会议室、客厅角落", "低光办公样例"],
    [101, "豆瓣绿", "peperomia_obtusifolia", "碧玉椒草", "/profile/upload/plant/1015.jpg", "小空间新手样例植物", 1, 1, "书桌、床头柜、卧室边柜", "桌面样例"],
    [104, "迷迭香", "rosmarinus_officinalis", "迷迭香草", "/profile/upload/plant/1017.jpg", "阳台闻香样例植物", 2, 2, "南向阳台、厨房窗边、露台", "闻香样例"],
    [102, "琴叶榕", "ficus_lyrata", "Fiddle Leaf Fig", "/profile/upload/plant/1018.jpg", "高预算陈设样例植物", 2, 3, "客厅落地、会客区、前台", "高预算陈设样例"],
    [102, "散尾葵", "dypsis_lutescens", "黄椰子", "/profile/upload/plant/1020.jpg", "宠物家庭客厅样例植物", 1, 2, "客厅、休闲区、办公室会客位", "宠物家庭样例"],
]

TAG_TEMPLATE_HEADERS = ["标签编码", "标签名称", "标签类型", "标签说明", "状态", "备注"]
TAG_TEMPLATE_ROWS = [
    ["light_low", "低光照", "light", "适合弱光或半阴环境", 0, "卧室与办公室高频标签"],
    ["pet_yes", "宠物友好", "pet", "宠物家庭优先推荐", 0, "安全性高频标签"],
    ["function_aroma", "芳香体验", "function", "适合阳台闻香与观花场景", 0, "闻香场景标签"],
    ["style_tropical", "热带感", "style", "适合客厅大叶氛围场景", 0, "风格标签"],
    ["feature_statement", "陈设型", "feature", "适合客厅主景和会客区焦点位", 0, "陈设标签"],
]


def ensure_dirs() -> None:
    for folder in (PLANT_DIR, ARTICLE_DIR, IMPORT_DIR):
        folder.mkdir(parents=True, exist_ok=True)


def load_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    font_candidates = []
    if bold:
        font_candidates.extend([
            Path("C:/Windows/Fonts/msyhbd.ttc"),
            Path("C:/Windows/Fonts/simhei.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ])
    font_candidates.extend([
        Path("C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simsun.ttc"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ])
    for candidate in font_candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size=size)
    return ImageFont.load_default()


def hex_to_rgb(color: str) -> tuple[int, int, int]:
    color = color.lstrip("#")
    return tuple(int(color[i:i + 2], 16) for i in (0, 2, 4))


def vertical_gradient(size: tuple[int, int], top_color: str, bottom_color: str) -> Image.Image:
    width, height = size
    top = hex_to_rgb(top_color)
    bottom = hex_to_rgb(bottom_color)
    image = Image.new("RGB", size, top)
    draw = ImageDraw.Draw(image)
    for y in range(height):
        ratio = y / max(height - 1, 1)
        color = tuple(int(top[index] + (bottom[index] - top[index]) * ratio) for index in range(3))
        draw.line((0, y, width, y), fill=color)
    return image


def add_soft_glow(canvas: Image.Image, color: str, box: tuple[int, int, int, int], blur_radius: int) -> None:
    layer = Image.new("RGBA", canvas.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(layer)
    draw.ellipse(box, fill=hex_to_rgb(color) + (110,))
    layer = layer.filter(ImageFilter.GaussianBlur(blur_radius))
    canvas.alpha_composite(layer)


def draw_leaf(canvas: Image.Image, center: tuple[int, int], size: tuple[int, int], fill: str, angle: float = 0.0, vein: str = "#d5ead4") -> None:
    leaf = Image.new("RGBA", (size[0] * 2, size[1] * 2), (0, 0, 0, 0))
    draw = ImageDraw.Draw(leaf)
    draw.ellipse((size[0] // 2, 0, size[0] + size[0] // 2, size[1] * 2 - 4), fill=hex_to_rgb(fill) + (255,))
    draw.line((size[0], 10, size[0], size[1] * 2 - 10), fill=hex_to_rgb(vein) + (150,), width=3)
    rotated = leaf.rotate(angle, resample=Image.Resampling.BICUBIC, expand=True)
    canvas.alpha_composite(rotated, (int(center[0] - rotated.width / 2), int(center[1] - rotated.height / 2)))


def draw_monstera_leaf(canvas: Image.Image, center: tuple[int, int], scale: float, fill: str) -> None:
    width = int(210 * scale)
    height = int(260 * scale)
    leaf = Image.new("RGBA", (width + 80, height + 80), (0, 0, 0, 0))
    draw = ImageDraw.Draw(leaf)
    polygon = [
        (width * 0.5, 10),
        (width * 0.9, height * 0.18),
        (width + 30, height * 0.5),
        (width * 0.85, height * 0.9),
        (width * 0.55, height + 20),
        (width * 0.2, height * 0.92),
        (20, height * 0.5),
        (width * 0.18, height * 0.18),
    ]
    draw.polygon(polygon, fill=hex_to_rgb(fill) + (255,))
    for x_offset in (0.32, 0.48, 0.64):
        slot_top = int(height * (0.28 + (x_offset - 0.32) * 0.2))
        slot_bottom = int(height * (0.64 + (x_offset - 0.32) * 0.15))
        draw.ellipse((int(width * x_offset), slot_top, int(width * x_offset + width * 0.12), slot_bottom), fill=(0, 0, 0, 0))
    draw.line((width * 0.52, 20, width * 0.52, height + 10), fill=(219, 236, 211, 200), width=5)
    rotated = leaf.rotate(-12, resample=Image.Resampling.BICUBIC, expand=True)
    canvas.alpha_composite(rotated, (int(center[0] - rotated.width / 2), int(center[1] - rotated.height / 2)))


def draw_pot(draw: ImageDraw.ImageDraw, x: int, y: int, width: int, height: int, fill: str, rim: str) -> None:
    draw.rounded_rectangle((x, y, x + width, y + height), radius=24, fill=hex_to_rgb(fill), outline=(0, 0, 0, 18), width=2)
    draw.rounded_rectangle((x - 8, y - 12, x + width + 8, y + 18), radius=18, fill=hex_to_rgb(rim))
    draw.rectangle((x + width * 0.25, y + height, x + width * 0.75, y + height + 16), fill=(120, 91, 58))


def add_room_backdrop(canvas: Image.Image, palette: tuple[str, str, str]) -> None:
    draw = ImageDraw.Draw(canvas)
    width, height = canvas.size
    draw.rectangle((0, height * 0.72, width, height), fill=(215, 196, 167))
    draw.rectangle((width * 0.08, height * 0.12, width * 0.44, height * 0.64), fill=(255, 255, 255, 70), outline=(255, 255, 255, 110), width=5)
    draw.rectangle((width * 0.12, height * 0.16, width * 0.40, height * 0.60), fill=(221, 240, 227))
    draw.line((width * 0.26, height * 0.16, width * 0.26, height * 0.60), fill=(255, 255, 255, 140), width=4)
    draw.line((width * 0.12, height * 0.38, width * 0.40, height * 0.38), fill=(255, 255, 255, 140), width=4)
    draw.rounded_rectangle((width * 0.62, height * 0.18, width * 0.86, height * 0.28), radius=24, fill=hex_to_rgb(palette[0]))
    draw.rectangle((width * 0.66, height * 0.28, width * 0.82, height * 0.32), fill=(126, 141, 98))
    draw.rectangle((width * 0.17, height * 0.68, width * 0.83, height * 0.72), fill=(121, 92, 67))


def draw_variant(canvas: Image.Image, variant: str, palette: tuple[str, str, str]) -> None:
    draw = ImageDraw.Draw(canvas)
    width, height = canvas.size
    pot_x = width // 2 - 120
    pot_y = height - 250
    draw_pot(draw, pot_x, pot_y, 240, 120, fill="#a76b41", rim="#c9895b")
    stem_color = "#5f472b"
    leaf_fill = palette[1]
    dark_fill = palette[2]

    if variant == "vine":
        draw.line((width * 0.5, pot_y, width * 0.5, 310), fill=hex_to_rgb(stem_color), width=10)
        vine_points = [(width * 0.5, 360), (width * 0.4, 420), (width * 0.32, 520), (width * 0.27, 620)]
        draw.line(vine_points, fill=hex_to_rgb(stem_color), width=7)
        for index, point in enumerate(vine_points[1:], start=1):
            draw_leaf(canvas, (int(point[0] + 36), int(point[1] - 20)), (68, 36), leaf_fill, angle=24 - index * 8)
            draw_leaf(canvas, (int(point[0] - 40), int(point[1] + 12)), (74, 40), dark_fill, angle=-35 + index * 10)
        for offset in range(6):
            draw_leaf(canvas, (width // 2 - 140 + offset * 56, 300 + (offset % 2) * 24), (86, 48), leaf_fill, angle=-35 + offset * 10)
    elif variant == "spider":
        for angle in (-66, -42, -18, 18, 42, 66):
            branch_end = (
                int(width / 2 + math.cos(math.radians(90 + angle)) * 210),
                int(pot_y - 35 + math.sin(math.radians(90 + angle)) * 210),
            )
            draw.line((width / 2, pot_y, branch_end[0], branch_end[1]), fill=hex_to_rgb(stem_color), width=5)
            draw_leaf(canvas, branch_end, (110, 24), leaf_fill, angle=angle)
            draw_leaf(canvas, (branch_end[0] + 10, branch_end[1] + 18), (92, 20), "#d8efbe", angle=angle)
    elif variant == "snake":
        for index in range(8):
            x = width * 0.28 + index * 56
            top = 180 + (index % 3) * 30
            draw.rounded_rectangle((x, top, x + 34, pot_y + 20), radius=18, fill=hex_to_rgb(leaf_fill), outline=hex_to_rgb(dark_fill), width=3)
            for stripe in range(5):
                y = top + 36 + stripe * 68
                draw.arc((x - 4, y, x + 38, y + 24), start=190, end=350, fill=(220, 230, 176), width=3)
    elif variant == "monstera":
        draw.line((width * 0.5, pot_y, width * 0.47, 410), fill=hex_to_rgb(stem_color), width=9)
        draw.line((width * 0.5, pot_y, width * 0.62, 470), fill=hex_to_rgb(stem_color), width=9)
        draw.line((width * 0.5, pot_y, width * 0.36, 500), fill=hex_to_rgb(stem_color), width=9)
        draw_monstera_leaf(canvas, (int(width * 0.63), 410), 0.82, leaf_fill)
        draw_monstera_leaf(canvas, (int(width * 0.38), 500), 0.75, dark_fill)
        draw_monstera_leaf(canvas, (int(width * 0.54), 280), 0.72, "#76b879")
    elif variant == "money_tree":
        for offset in (-30, 0, 30):
            draw.line((width * 0.5 + offset, pot_y, width * 0.5 + offset * 0.35, 360), fill=hex_to_rgb(stem_color), width=12)
        for branch in ((-120, 290), (-30, 250), (60, 280), (140, 330)):
            for angle in (-32, 0, 32):
                draw_leaf(canvas, (int(width * 0.5 + branch[0] + angle), branch[1] + abs(angle) // 4), (96, 42), leaf_fill, angle=angle)
    elif variant == "peace_lily":
        for offset in (-120, -56, 0, 56, 120):
            draw.line((width * 0.5, pot_y, width * 0.5 + offset * 0.45, 360 + abs(offset) // 4), fill=hex_to_rgb(stem_color), width=7)
            draw_leaf(canvas, (int(width * 0.5 + offset * 0.55), 360 + abs(offset) // 5), (92, 42), leaf_fill, angle=offset * 0.18)
        for flower_x in (-80, 10, 100):
            draw.line((width * 0.5 + flower_x * 0.3, pot_y, width * 0.5 + flower_x, 260 + abs(flower_x) // 3), fill=hex_to_rgb(stem_color), width=4)
            draw.ellipse((width * 0.5 + flower_x - 32, 240 + abs(flower_x) // 3, width * 0.5 + flower_x + 28, 320 + abs(flower_x) // 3), fill=(248, 245, 235), outline=(226, 220, 210), width=2)
            draw.rectangle((width * 0.5 + flower_x - 3, 274 + abs(flower_x) // 3, width * 0.5 + flower_x + 3, 316 + abs(flower_x) // 3), fill=(231, 196, 77))
    elif variant == "ivy":
        draw.line((width * 0.5, pot_y, width * 0.5, 350), fill=hex_to_rgb(stem_color), width=8)
        for path in (
            [(width * 0.48, 360), (width * 0.38, 430), (width * 0.28, 510), (width * 0.22, 620)],
            [(width * 0.54, 360), (width * 0.64, 440), (width * 0.72, 520), (width * 0.78, 620)],
        ):
            draw.line(path, fill=hex_to_rgb(stem_color), width=5)
            for point in path[1:]:
                draw_leaf(canvas, (int(point[0]), int(point[1])), (72, 34), leaf_fill, angle=-18)
                draw_leaf(canvas, (int(point[0] + 30), int(point[1] - 8)), (62, 30), dark_fill, angle=16)
    elif variant == "rubber":
        draw.line((width * 0.5, pot_y, width * 0.5, 300), fill=hex_to_rgb(stem_color), width=11)
        for x_offset, y_offset, angle in ((-140, 420, -52), (-92, 330, -24), (96, 330, 24), (152, 420, 52), (0, 250, 0)):
            draw_leaf(canvas, (int(width * 0.5 + x_offset), y_offset), (120, 60), dark_fill if abs(x_offset) > 100 else leaf_fill, angle=angle, vein="#f1f2d4")
    elif variant == "fern":
        for branch in range(-4, 5):
            branch_x = width * 0.5 + branch * 42
            top_y = 260 + abs(branch) * 22
            draw.line((width * 0.5, pot_y, branch_x, top_y), fill=hex_to_rgb(stem_color), width=4)
            for leaf_index in range(7):
                y = top_y + leaf_index * 40
                draw_leaf(canvas, (int(branch_x - 22), y), (40, 16), leaf_fill, angle=-55)
                draw_leaf(canvas, (int(branch_x + 22), y + 10), (40, 16), dark_fill, angle=55)
    elif variant == "jasmine":
        for offset in (-70, 0, 70):
            draw.line((width * 0.5, pot_y, width * 0.5 + offset, 340 + abs(offset) // 3), fill=hex_to_rgb(stem_color), width=5)
        for flower in ((width * 0.42, 300), (width * 0.55, 260), (width * 0.66, 330), (width * 0.34, 370)):
            for petal_angle in range(0, 360, 72):
                dx = math.cos(math.radians(petal_angle)) * 22
                dy = math.sin(math.radians(petal_angle)) * 22
                draw.ellipse((flower[0] + dx - 16, flower[1] + dy - 12, flower[0] + dx + 16, flower[1] + dy + 12), fill=(252, 249, 240), outline=(232, 226, 212))
            draw.ellipse((flower[0] - 10, flower[1] - 10, flower[0] + 10, flower[1] + 10), fill=(234, 196, 77))
            draw_leaf(canvas, (int(flower[0] - 42), int(flower[1] + 34)), (62, 28), leaf_fill, angle=-28)
            draw_leaf(canvas, (int(flower[0] + 40), int(flower[1] + 20)), (62, 28), dark_fill, angle=22)
    elif variant == "mint":
        for x_offset in (-90, -36, 24, 86):
            draw.line((width * 0.5 + x_offset * 0.25, pot_y, width * 0.5 + x_offset, 320 + abs(x_offset) // 4), fill=hex_to_rgb(stem_color), width=5)
            for y in range(360, pot_y, 62):
                draw_leaf(canvas, (int(width * 0.5 + x_offset - 22), y), (56, 22), leaf_fill, angle=-35)
                draw_leaf(canvas, (int(width * 0.5 + x_offset + 22), y + 10), (56, 22), dark_fill, angle=35)
    elif variant == "succulent":
        for ring, count in ((0, 8), (1, 10), (2, 12)):
            radius = 36 + ring * 34
            size = (90 - ring * 16, 36 - ring * 6)
            for index in range(count):
                angle = (360 / count) * index
                x = width * 0.5 + math.cos(math.radians(angle)) * radius
                y = pot_y - 14 + math.sin(math.radians(angle)) * radius * 0.55
                fill = leaf_fill if index % 2 == 0 else dark_fill
                draw_leaf(canvas, (int(x), int(y)), (int(size[0]), int(size[1])), fill, angle=angle)
    elif variant == "zz":
        for x_offset in (-110, -46, 22, 92):
            stem_top = 280 + abs(x_offset) // 5
            draw.line((width * 0.5 + x_offset * 0.22, pot_y, width * 0.5 + x_offset, stem_top), fill=hex_to_rgb(stem_color), width=6)
            for step in range(5):
                y = stem_top + step * 44
                draw_leaf(canvas, (int(width * 0.5 + x_offset - 22), y), (56, 22), leaf_fill, angle=-34)
                draw_leaf(canvas, (int(width * 0.5 + x_offset + 22), y + 10), (56, 22), dark_fill, angle=34)
    elif variant == "palm":
        for angle in (-62, -36, -12, 12, 36, 62):
            end_x = int(width * 0.5 + math.sin(math.radians(angle)) * 230)
            end_y = int(pot_y - 130 - math.cos(math.radians(angle)) * 90)
            draw.line((width * 0.5, pot_y, end_x, end_y), fill=hex_to_rgb(stem_color), width=5)
            for leaflet in range(6):
                ratio = leaflet / 5
                x = int(width * 0.5 + (end_x - width * 0.5) * ratio)
                y = int(pot_y + (end_y - pot_y) * ratio)
                draw_leaf(canvas, (x - 16, y - 6), (44, 14), leaf_fill, angle=angle - 52)
                draw_leaf(canvas, (x + 18, y + 4), (44, 14), dark_fill, angle=angle + 52)
    elif variant == "peperomia":
        for x_offset, height_offset in ((-90, 360), (-40, 300), (18, 330), (84, 288)):
            draw.line((width * 0.5 + x_offset * 0.25, pot_y, width * 0.5 + x_offset, height_offset), fill=hex_to_rgb(stem_color), width=5)
            for leaf_offset in (-26, 0, 26):
                draw.ellipse((width * 0.5 + x_offset + leaf_offset - 28, height_offset - 26 + abs(leaf_offset) // 5, width * 0.5 + x_offset + leaf_offset + 28, height_offset + 26 + abs(leaf_offset) // 5), fill=hex_to_rgb(leaf_fill if leaf_offset != 0 else dark_fill) + (255,), outline=(232, 242, 218), width=2)
    elif variant == "rosemary":
        for x_offset in (-70, -26, 20, 66):
            end_y = 260 + abs(x_offset) // 3
            draw.line((width * 0.5 + x_offset * 0.2, pot_y, width * 0.5 + x_offset, end_y), fill=hex_to_rgb(stem_color), width=4)
            for step in range(9):
                y = end_y + step * 34
                draw.line((width * 0.5 + x_offset - 18, y, width * 0.5 + x_offset + 2, y - 10), fill=hex_to_rgb(leaf_fill), width=3)
                draw.line((width * 0.5 + x_offset + 18, y + 4, width * 0.5 + x_offset - 2, y - 8), fill=hex_to_rgb(dark_fill), width=3)
    elif variant == "fiddle":
        for x_offset in (-34, 0, 34):
            draw.line((width * 0.5 + x_offset * 0.2, pot_y, width * 0.5 + x_offset * 0.5, 330 + abs(x_offset)), fill=hex_to_rgb(stem_color), width=8)
        for center_x, center_y, angle in ((width * 0.38, 420, -36), (width * 0.47, 300, -12), (width * 0.58, 284, 18), (width * 0.66, 420, 38)):
            leaf = Image.new("RGBA", (220, 300), (0, 0, 0, 0))
            leaf_draw = ImageDraw.Draw(leaf)
            leaf_draw.ellipse((52, 20, 168, 138), fill=hex_to_rgb(leaf_fill) + (255,))
            leaf_draw.ellipse((40, 112, 180, 290), fill=hex_to_rgb(dark_fill) + (255,))
            leaf_draw.rectangle((86, 112, 134, 180), fill=hex_to_rgb(leaf_fill) + (255,))
            leaf_draw.line((110, 28, 110, 274), fill=(226, 238, 214, 170), width=5)
            rotated = leaf.rotate(angle, resample=Image.Resampling.BICUBIC, expand=True)
            canvas.alpha_composite(rotated, (int(center_x - rotated.width / 2), int(center_y - rotated.height / 2)))


def add_caption(canvas: Image.Image, title: str, subtitle: str) -> None:
    overlay = Image.new("RGBA", canvas.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    width, height = canvas.size
    draw.rounded_rectangle((70, height - 210, width - 70, height - 70), radius=28, fill=(255, 252, 246, 196))
    draw.text((110, height - 180), title, fill="#21351c", font=load_font(46, bold=True))
    draw.text((114, height - 116), subtitle, fill="#52624d", font=load_font(24))
    canvas.alpha_composite(overlay)


def generate_plant_image(asset: dict) -> str:
    image = vertical_gradient((1200, 900), asset["palette"][0], "#f7f2e6").convert("RGBA")
    add_soft_glow(image, asset["palette"][1], (720, 40, 1120, 360), 70)
    add_soft_glow(image, asset["palette"][0], (40, 90, 460, 460), 90)
    add_room_backdrop(image, asset["palette"])
    draw_variant(image, asset["variant"], asset["palette"])
    add_caption(image, asset["name"], asset["latin"])
    output_path = PLANT_DIR / f"{asset['plant_id']}.jpg"
    image.convert("RGB").save(output_path, quality=92)
    return f"/profile/upload/plant/{asset['plant_id']}.jpg"


def draw_scene_card(canvas: Image.Image, scene: str, palette: tuple[str, str, str]) -> None:
    draw = ImageDraw.Draw(canvas)
    width, height = canvas.size
    draw.rounded_rectangle((70, 70, width - 70, height - 70), radius=38, fill=(255, 255, 255, 72), outline=(255, 255, 255, 110), width=3)
    if scene == "bedroom":
        draw.rectangle((150, 380, 610, 540), fill=(217, 206, 188))
        draw.rectangle((120, 330, 640, 395), fill=(188, 164, 141))
        draw.rectangle((700, 160, 1100, 420), fill=(223, 237, 228))
        draw.rectangle((734, 194, 1066, 386), fill=(244, 248, 251))
        draw.line((900, 194, 900, 386), fill=(215, 228, 223), width=5)
        draw.line((734, 290, 1066, 290), fill=(215, 228, 223), width=5)
    elif scene == "office":
        draw.rectangle((140, 430, 1130, 500), fill=(187, 164, 134))
        draw.rectangle((220, 240, 520, 430), fill=(236, 243, 244))
        draw.rectangle((250, 270, 490, 400), fill=(161, 190, 200))
        draw.rectangle((690, 190, 990, 410), fill=(233, 239, 230))
        draw.rectangle((1020, 220, 1120, 430), fill=(202, 213, 205))
    elif scene == "balcony":
        draw.rectangle((720, 110, 1110, 420), fill=(223, 240, 230))
        draw.rectangle((760, 150, 1070, 380), fill=(182, 219, 230))
        draw.line((915, 150, 915, 380), fill=(237, 247, 242), width=5)
        draw.line((760, 265, 1070, 265), fill=(237, 247, 242), width=5)
        draw.rectangle((150, 470, 1110, 530), fill=(192, 162, 125))
        draw.ellipse((190, 180, 470, 470), fill=(255, 229, 163, 75))
    elif scene == "pet":
        draw.rectangle((180, 390, 560, 540), fill=(208, 199, 186))
        draw.rectangle((240, 270, 500, 390), fill=(232, 225, 214))
        draw.ellipse((760, 310, 1040, 520), fill=(247, 244, 239))
        draw.ellipse((806, 270, 888, 340), fill=(247, 244, 239))
        draw.ellipse((912, 270, 994, 340), fill=(247, 244, 239))
        draw.polygon([(816, 278), (852, 220), (876, 288)], fill=(247, 244, 239))
        draw.polygon([(922, 288), (946, 220), (982, 278)], fill=(247, 244, 239))
        draw.ellipse((860, 400, 884, 416), fill=(214, 181, 158))
        draw.line((848, 438, 896, 438), fill=(161, 122, 102), width=3)
    elif scene == "office_lowlight":
        draw.rectangle((150, 430, 1120, 500), fill=(173, 156, 135))
        draw.rectangle((250, 220, 500, 430), fill=(227, 232, 228))
        draw.rectangle((280, 250, 470, 395), fill=(138, 154, 156))
        draw.rectangle((760, 170, 1080, 390), fill=(196, 207, 198))
        draw.ellipse((890, 110, 990, 210), fill=(251, 229, 180, 90))
    elif scene == "living":
        draw.rectangle((150, 400, 1120, 530), fill=(194, 165, 135))
        draw.rounded_rectangle((180, 250, 620, 430), radius=36, fill=(216, 205, 193))
        draw.rounded_rectangle((700, 230, 1080, 430), radius=28, fill=(238, 232, 223))
        draw.ellipse((790, 275, 980, 455), fill=(228, 232, 222))
    elif scene == "desktop":
        draw.rectangle((140, 440, 1120, 510), fill=(182, 155, 125))
        draw.rectangle((250, 260, 620, 438), fill=(239, 243, 239))
        draw.rectangle((292, 292, 578, 402), fill=(181, 199, 192))
        draw.line((436, 438, 436, 500), fill=(126, 133, 118), width=6)
        draw.rectangle((760, 250, 1040, 390), fill=(242, 236, 228))
        draw.line((792, 282, 1008, 282), fill=(213, 200, 187), width=6)
        draw.line((792, 322, 968, 322), fill=(213, 200, 187), width=6)
    elif scene == "aroma":
        draw.rectangle((160, 470, 1110, 530), fill=(194, 165, 126))
        draw.rectangle((720, 120, 1090, 400), fill=(230, 240, 228))
        draw.rectangle((760, 160, 1050, 360), fill=(194, 223, 232))
        draw.line((905, 160, 905, 360), fill=(237, 247, 242), width=5)
        draw.line((760, 260, 1050, 260), fill=(237, 247, 242), width=5)
        draw.ellipse((210, 180, 520, 460), fill=(255, 229, 163, 68))
        draw.ellipse((250, 350, 330, 430), fill=(245, 244, 238))
        draw.ellipse((350, 350, 430, 430), fill=(245, 244, 238))
    draw_pot(draw, 850 if scene != "pet" else 575, 390, 170, 104, fill="#9d6a42", rim="#c78a5a")
    draw_leaf(canvas, (910 if scene != "pet" else 660, 320), (130, 54), palette[1], angle=-20)
    draw_leaf(canvas, (980 if scene != "pet" else 722, 370), (112, 48), palette[2], angle=22)
    draw_leaf(canvas, (840 if scene != "pet" else 612, 380), (102, 42), "#88bb78", angle=-52)


def generate_article_image(asset: dict) -> str:
    image = vertical_gradient((1280, 720), asset["palette"][0], "#f7f0e7").convert("RGBA")
    add_soft_glow(image, asset["palette"][1], (820, 20, 1220, 320), 80)
    add_soft_glow(image, asset["palette"][2], (50, 160, 340, 560), 100)
    draw_scene_card(image, asset["scene"], asset["palette"])
    add_caption(image, asset["title"], asset["subtitle"])
    output_path = ARTICLE_DIR / f"{asset['slug']}.jpg"
    image.convert("RGB").save(output_path, quality=92)
    return f"/profile/upload/article/{asset['slug']}.jpg"


def style_sheet(sheet, title: str, headers: list[str], rows: list[list[object]]) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="2F5C3A")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D7E2D4"),
        right=Side(style="thin", color="D7E2D4"),
        top=Side(style="thin", color="D7E2D4"),
        bottom=Side(style="thin", color="D7E2D4"),
    )
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=15, bold=True, color="21351C")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        sheet.column_dimensions[cell.column_letter].width = 18
    for row_index, row_data in enumerate(rows, start=3):
        for col_index, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
    sheet.freeze_panes = "A3"


def generate_template(file_name: str, title: str, headers: list[str], rows: list[list[object]]) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "demo"
    style_sheet(sheet, title, headers, rows)
    output_path = IMPORT_DIR / file_name
    workbook.save(output_path)
    return f"/profile/upload/import/{file_name}"


def write_manifest(plant_map: list[dict], article_map: list[dict], import_map: list[dict]) -> None:
    manifest = {
        "generatedAt": "deterministic-script",
        "runtimeUploadPath": "/profile/upload",
        "plantAssets": plant_map,
        "articleAssets": article_map,
        "importAssets": import_map,
    }
    MANIFEST_PATH.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    ensure_dirs()
    plant_map = []
    article_map = []
    import_map = []
    for asset in PLANT_ASSETS:
        resource_path = generate_plant_image(asset)
        plant_map.append({"plantId": asset["plant_id"], "plantName": asset["name"], "path": resource_path})
    for asset in ARTICLE_ASSETS:
        resource_path = generate_article_image(asset)
        article_map.append({"slug": asset["slug"], "title": asset["title"], "path": resource_path})
    import_map.append({
        "importType": "plant_basic",
        "path": generate_template("plant_basic_demo.xlsx", "植物主数据导入模板", PLANT_TEMPLATE_HEADERS, PLANT_TEMPLATE_ROWS),
    })
    import_map.append({
        "importType": "tag_basic",
        "path": generate_template("plant_tag_demo.xlsx", "植物标签导入模板", TAG_TEMPLATE_HEADERS, TAG_TEMPLATE_ROWS),
    })
    write_manifest(plant_map, article_map, import_map)
    print(f"Generated {len(plant_map)} plant images, {len(article_map)} article images and {len(import_map)} templates into {UPLOAD_ROOT}")


if __name__ == "__main__":
    main()